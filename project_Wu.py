
# coding: utf-8

import os, struct
from openpyxl import load_workbook
import numpy as np
import numpy.linalg as LA
from numpy import *  
import matplotlib  
import matplotlib.pyplot as plt  


# In[1]
#get the training data and store them in a list——mat1
#Dataset in totall 32561 rows，11 columns
def get_data1():
    data_=[[0 for col in range(11)] for row in range(32561)]
    filename='C:/Users/wuxia/Desktop/courses_8/ML/project/new_project_data.xlsx' 
    workbook_ = load_workbook(filename) #导入工作表
    sheetnames =workbook_.get_sheet_names() #获得表单名字
    print(sheetnames)
   
    sheet = workbook_.get_sheet_by_name('data') #从工作表中提取某一表单
    for rowNum in range(2,32563):
        for colNum in range(1,12):
           data_[rowNum-2][colNum-1]=sheet.cell(row=rowNum,column=colNum).value

    return(data_) 

mat1=get_data1()


# In[2]
# build lists for each features
age=[]  
work=[] 
education=[]
marital=[]
job=[]
relation=[]
race=[]
sex=[]
hour=[]
nation=[]
income=[]

for i in range(32561):
    age.append(mat1[i][0])
    work.append(mat1[i][1])
    education.append(mat1[i][2])
    marital.append(mat1[i][3])
    job.append(mat1[i][4])
    relation.append(mat1[i][5])
    race.append((mat1[i][6]))
    sex.append((mat1[i][7]))
    hour.append(mat1[i][8])
    nation.append(mat1[i][9])
    income.append(mat1[i][10])
    
# the values that each feature can have are listed below
work2=[' ?', ' Federal-gov', ' Private', ' Without-pay', ' State-gov', ' Local-gov', ' Never-worked', ' Self-emp-inc', ' Self-emp-not-inc']
marital2=[' Divorced', ' Never-married', ' Married-civ-spouse', ' Widowed', ' Separated', ' Married-spouse-absent', ' Married-AF-spouse']
job2=[' ?', ' Farming-fishing', ' Adm-clerical', ' Prof-specialty', ' Priv-house-serv', ' Sales', ' Craft-repair', ' Transport-moving', ' Armed-Forces', ' Protective-serv', ' Machine-op-inspct', ' Tech-support', ' Exec-managerial', ' Other-service', ' Handlers-cleaners']
relation2=[' Husband', ' Wife', ' Not-in-family', ' Other-relative', ' Unmarried', ' Own-child']
race2=[' Amer-Indian-Eskimo', ' Asian-Pac-Islander', ' Black', ' Other', ' White']
sex2=[' Female', ' Male']
nation2=[' Germany', ' France', ' Nicaragua', ' Iran', ' China', ' Guatemala', ' ?', ' Honduras', ' Dominican-Republic', ' Greece', ' United-States', ' Hungary', ' Jamaica', ' Peru', ' Laos', ' Puerto-Rico', ' Outlying-US(Guam-USVI-etc)', ' Scotland', ' Canada', ' Trinadad&Tobago', ' Vietnam', ' Cambodia', ' Italy', ' Hong', ' Philippines', ' Yugoslavia', ' Columbia', ' Taiwan', ' Ecuador', ' Ireland', ' South', ' Portugal', ' Japan', ' Cuba', ' Mexico', ' El-Salvador', ' Holand-Netherlands', ' England', ' India', ' Poland', ' Haiti', ' Thailand']
income2=[' <=50K', ' >50K']



# In[3]
# Use Kesler's method to convert nominal values to numerical values
# We get feature_mat and lable_mat
def Kesler(mat,mat2):
    n=len(mat2)
    mat3=np.zeros([32561,n])
    for i in range(32561):
        for j in range(n):
            if mat[i]==mat2[j]:
               mat3[i][j]=1
            else:
               mat3[i][j]=-1
    
    return(mat3)

# After the Kesler's method, the new values are listed below
K_age=np.matrix(age).T
K_work=np.matrix(Kesler(work,work2))
K_education=np.matrix(education).T
K_marital=np.matrix(Kesler(marital,marital2))
K_job=np.matrix(Kesler(job,job2))
K_relation=np.matrix(Kesler(relation,relation2))
K_race=np.matrix(Kesler(race,race2))
K_sex=np.zeros([32561,1])
K_hour=np.matrix(hour).T
K_nation=np.matrix(Kesler(nation,nation2))
K_income=np.zeros([32561,1])

for i in range(32561):
    if sex[i]==sex2[1]:
        K_sex[i]=1  # male's representated by 1
    else:
        K_sex[i]=-1 #female's representated by -1
             
    if income[i]==income2[1]:
        K_income[i]=1 #income higher than 50k denoted by 1
    else:
        K_income[i]=-1 #income lower than 50k denoted by -1
    
#the dimension for the feature vectors are 88
feature_mat=np.concatenate((K_age, K_work,K_education,K_marital,K_job,K_relation,K_race,K_sex,K_hour,K_nation), axis=1)
lable_mat=K_income



# In[4]
# Let's build a linear classifier
Xa=np.concatenate((np.ones([32561,1]),feature_mat), axis=1)
W=np.dot(np.linalg.pinv(Xa),lable_mat)

#Let's test it!
result1=np.dot(Xa,W)
for i in range(len(result1)):
    if result1[i]>0:
        result1[i]=1
    elif result1[i]<0:
        result1[i]=-1
        
FF,FP,PF,PP=0,0,0,0
for i in range(len(lable_mat)):
    if (lable_mat[i]==-1 and result1[i]==-1):
        FF=FF+1
    elif (lable_mat[i]==-1 and result1[i]==1):
        FP=FP+1
    elif (lable_mat[i]==1 and result1[i]==-1):
        PF=PF+1
    elif (lable_mat[i]==1 and result1[i]==1):
        PP=PP+1
    else:
        print('an error occurs in[5]')
print('FF=',FF)
print('FP=',FP)
print('PF=',PF)
print('PP=',PP)        
print('accuracy1=',(FF+PP)/(FF+FP+PF+PP))
print('Sensitivity1=',FF/(FF+FP))
print('specificity1=',PP/(PF+PP))
print('PPV=',FF/(FF+PF))



# In[5]:
#let's do the PCA!

#X is feature_mat
μ=np.mean(feature_mat,axis=0);

#get Z and 
Z=feature_mat-μ

#get C
C=np.cov(Z,rowvar=False);

#get V
#the columns of V are the eigenvectors
[λ,V]=LA.eigh(C);
# For notational agreement with the Lecture Notes, we will set V to its own transpose
# Now, the rows are eigenvectors
λ=np.flipud(λ); V=np.flipud(V.T);
v1=V[0,:] #eigen1
v2=V[1,:] #eigen2
         
         
#get P
P=np.dot(Z,V.T);

#R
R=np.dot(P,V);

#check Xrec
Xrec=R+μ;diff=Xrec-feature_mat



# In[6]
#number of positive(>50K) and negative numbers(<=50K)
Np=0 
Nn=0
for i in range(len(lable_mat)):
    if lable_mat[i]==-1:
     Nn=Nn+1
    else:
     Np=Np+1

print('Number of negative samples',Nn)
print('Number of positive samples',Np)

#divide the data into 2 groups
#Xp(+1), Xn(-1)
Xp=np.zeros([Np,88]);
Xn=np.zeros([Nn,88]);
k=-1
j=-1
for i in range(len(lable_mat)):
    if lable_mat[i]==-1:
     k=k+1
     Xn[k]=feature_mat[i]
    else:
     j=j+1
     Xp[j]=feature_mat[i]


#scatter plot
Pp=np.zeros([Np,2])
Pn=np.zeros([Nn,2])
k=-1
j=-1
for i in range(len(lable_mat)):
    if lable_mat[i]==-1:
     k=k+1
     Pn[k][0]=P[i,0]
     Pn[k][1]=P[i,1]
    else:
     j=j+1
     Pp[j][0]=P[i,0]
     Pp[j][1]=P[i,1]


#p1 = plt.scatter(Pp[:,0], Pp[:,1], marker = '.', color = 'r') #red is positive
#p2 = plt.scatter(Pn[:,0], Pn[:,1], marker = '.', color = 'b') #blue is negative

plt.scatter(Pn[:,0], Pn[:,1], marker = '.', color = 'b') #blue is negative
plt.scatter(Pp[:,0], Pp[:,1], marker = '.', color = 'r') #red is positive






